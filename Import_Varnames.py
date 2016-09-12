# -*- coding: utf-8 -*-
"""
Import data from the Excel file in which model parameters are stored. Creates an object
called "estimates", which contains all of the model inputs in the following form:

    estimates.varname == 'name': <Estimate: type, mean, standard error>

IMPORTANT: If you want to use this file, you will have to make sure the Excel file is in the
            same directory as this file, or change the "load_workbook" path to the appropriate
            directory.

Variable Type:
    1 - Beta distributed (value between 0 and 1)
    2 - Normally distributed (value between negative and positive infinity)
    3 - Time to Event (value between 0 and infinity)
    4 - Gamma distributed (value between 0 and infinity)
    5 - Dirichlet distributed (create >2 probabilities that sum to 1.0 based on counts)
    
Dirichlet-distributed variables must be 
    
Code provided by Stavros Korokithakis of Stochastic Technologies (www.stavros.io)

"""

from openpyxl import load_workbook                  # Load the import function
import numpy

class Estimates:                                    # An empty class to hold data
    pass

class Estimate:                                     # A class to process the data
    def __init__(self, etype, mean, se):
        self.type = etype                           # Define variable type
        self.mean = mean                            # Define mean value
        self.se = se                                # Define standard error

    def sample(self):                               # A function that checks variable type and samples a value accordingly
        if self.type == 1:                                              # Beta-distributed variables (probabilities)
            x = self.mean
            y = self.se
            bdist_alpha = x*((x*(1-x)/y**2) - 1)                        # A formula to produce the alpha parameter
            bdist_beta = (1-x)*(x/y**2*(1-x) - 1)                       # A formula to produce the beta parameter
            samp_value = numpy.random.beta(bdist_alpha, bdist_beta)
            return samp_value
            
        elif self.type == 2:                                            # Normally-distributed variables
            samp_value = numpy.random.normal(self.mean, self.se)
            return samp_value
            
        elif self.type == 3:
            samp_value = numpy.random.normal(self.mean, self.se)   # Aug 4 2016 - THIS IS WRONG AND I NEED TO FIGURE OUT WHAT TO DO WITH TTE DATA
            return samp_value

        elif self.type == 4:                                            # Gamma-distributed variables
            x = self.mean
            y = self.se
            gdist_alpha <- x**2/y**2                                    # A formula to produce the shape parameter
            gdist_beta <- y**2/x                                        # A formula to produce the scale parameter
            
        elif self.type == 5:
            print("Variables of this type should use the 'dirisample' function")
            
        else:
            print("Estimates for this variable must be produced directly")


    def __str__(self):
        return "<Estimate: %s, %s, %s>" % (self.type, self.mean, self.se)
    __repr__ = __str__
    
def diriSample(estimates, names, values):           # A function that produces 'broken stick' estimates of probabilities from counts of >2 inputs
    dirich = numpy.random.dirichlet(values)
    for i in range(0,len(values)):
        setattr(estimates, names[i], dirich[i])     # Produce an estimate of each variable, and assign it the name from the "names" string
    

workbook = load_workbook('ImportTest.xlsx')
sheet = workbook.active

estimates = Estimates()

for line in sheet.rows:
    if not line[0].value:
        # There's no estimate name in this row.
        continue
    setattr(estimates, line[0].value, Estimate(line[1].value, line[2].value, line[3].value))

del(estimates.Parameter)

# To generate a sampled value from this function, the syntax is "estimates.varname.sample()"